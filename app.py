# --------------------------------------------------------------------------
#                           라이브러리 불러오기
# --------------------------------------------------------------------------
import streamlit as st
import pandas as pd
import plotly.express as px
import numpy as np
import io
import os
import csv
from datetime import datetime
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
#                           1. 기본 페이지 설정
# --------------------------------------------------------------------------
st.set_page_config(page_title="비급여 항목 비교 분석", layout="wide")
st.title("비급여 항목 비교 분석")

# --------------------------------------------------------------------------
#                           2. 데이터 로딩 함수
# --------------------------------------------------------------------------
@st.cache_data
def load_data(file_path):
    try:
        df = pd.read_parquet(file_path)
        if 'npay_code' in df.columns:
            df['npay_code'] = df['npay_code'].astype(str)
        return df
    except FileNotFoundError:
        return None

# --------------------------------------------------------------------------
#                           3. 엑셀 리포트 생성 함수
# --------------------------------------------------------------------------
def create_excel_report(item_name, stats, median_price, hospital_data):
    output = io.BytesIO()
    hospital_data_excel = hospital_data.copy()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_data = {
            "항목": ["분석 항목명", "평균 가격", "중앙값", "최저가", "최고가", "취급 병원 수"],
            "값": [
                item_name,
                f"{int(stats['mean']):,} 원",
                f"{int(median_price):,} 원",
                f"{int(stats['min']):,} 원",
                f"{int(stats['max']):,} 원",
                f"{int(stats['count'])} 곳"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='분석 리포트', index=False, startrow=0)
        hospital_data_excel.to_excel(writer, sheet_name='분석 리포트', index=False, startrow=len(summary_df) + 2)
        
        worksheet = writer.sheets['분석 리포트']
        try:
            price_col_idx = hospital_data_excel.columns.get_loc('price')
            price_col_letter = get_column_letter(price_col_idx + 1)
            start_row = len(summary_df) + 4
            for i in range(start_row, start_row + len(hospital_data_excel)):
                worksheet[f'{price_col_letter}{i}'].number_format = '#,##0'
        except KeyError:
            pass

        for i, col in enumerate(hospital_data_excel.columns):
            column_letter = get_column_letter(i + 1)
            column_len = max(hospital_data_excel[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[column_letter].width = column_len

    return output.getvalue()

# --------------------------------------------------------------------------
#                           4. 공통 분석 페이지 생성 함수
# --------------------------------------------------------------------------
def create_analysis_page(df):
    # --- 사이드바 UI ---
    st.sidebar.header("검색")
    search_options = ['항목명', '병원명', '항목 코드']
    selected_scopes = st.sidebar.multiselect("검색할 범위를 선택하세요", options=search_options, default=['항목명'], key=f"scopes_{df.hashCode}")
    search_keyword = st.sidebar.text_input("검색어를 입력하세요", placeholder="선택된 범위 내에서 검색합니다", key=f"search_{df.hashCode}")

    st.sidebar.divider()
    st.sidebar.header("상세 분석")
    item_search_keyword = st.sidebar.text_input("분석할 항목 검색", placeholder="분석하고 싶은 항목을 검색하세요", key=f"item_search_{df.hashCode}")

    if item_search_keyword:
        filtered_item_list = sorted([item for item in df['item_name'].unique() if item_search_keyword.lower() in item.lower()])
    else:
        filtered_item_list = []
    selected_item = st.sidebar.selectbox("항목 선택 (위에서 먼저 검색)", options=filtered_item_list, index=None, placeholder="항목을 선택하세요", key=f"item_select_{df.hashCode}")

    # --- 메인 화면 UI ---
    st.header("데이터 조회 결과")
    if search_keyword and selected_scopes:
        conditions = []
        if '항목명' in selected_scopes: conditions.append(df['item_name'].str.contains(search_keyword, case=False, na=False))
        if '병원명' in selected_scopes: conditions.append(df['hospital_name'].str.contains(search_keyword, case=False, na=False))
        if '항목 코드' in selected_scopes and 'npay_code' in df.columns: conditions.append(df['npay_code'].str.contains(search_keyword, case=False, na=False))
        if conditions:
            final_condition = np.logical_or.reduce(conditions)
            df_filtered = df[final_condition]
        else: df_filtered = df
    elif not selected_scopes and search_keyword:
        st.warning("검색할 범위를 1개 이상 선택해주세요.")
        df_filtered = pd.DataFrame()
    else: df_filtered = df
    st.dataframe(df_filtered)
    st.info(f"조회된 항목 수: **{len(df_filtered)}** 건")

    if selected_item:
        st.divider()
        st.header(f"'{selected_item}' 상세 분석")
        df_selected = df[df['item_name'] == selected_item]
        stats = df_selected['price'].describe()
        median_price = df_selected['price'].median()
        
        st.subheader("통계 요약")
        col1, col2, col3, col4, col5 = st.columns(5)
        col1.metric("평균 가격", f"{int(stats['mean']):,} 원")
        col2.metric("중앙값", f"{int(median_price):,} 원")
        col3.metric("최저가", f"{int(stats['min']):,} 원")
        col4.metric("최고가", f"{int(stats['max']):,} 원")
        col5.metric("취급 병원 수", f"{int(stats['count'])} 곳")

        st.subheader("우리 병원 가격 순위 (삼성서울병원)")
        our_hospital_name = "삼성서울병원"
        df_ranked = df_selected.sort_values(by='price', ascending=False).reset_index(drop=True)
        df_ranked['rank'] = df_ranked.index + 1
        hospital_rank_info = df_ranked[df_ranked['hospital_name'] == our_hospital_name]

        if not hospital_rank_info.empty:
            rank = hospital_rank_info['rank'].iloc[0]
            price = hospital_rank_info['price'].iloc[0]
            total_hospitals = len(df_ranked)
            st.metric(label=f"{our_hospital_name} 순위", value=f"{rank} 위 / {total_hospitals} 곳")
            st.text(f"가격: {price:,.0f} 원")
        else:
            st.warning(f"'{selected_item}' 항목에 대한 삼성서울병원 데이터를 찾을 수 없습니다.")
        
        st.subheader("병원별 가격 분포")
        fig = px.box(df_selected, x='hospital_name', y='price', title=f"'{selected_item}' 병원별 가격 분포", labels={'hospital_name': '병원명', 'price': '가격 (원)'})
        fig.update_xaxes(tickangle=45)
        st.plotly_chart(fig, use_container_width=True)

        st.divider()
        st.header("분석 리포트")
        st.subheader("전체 데이터 목록")
        report_df = df_selected.sort_values(by='price', ascending=False).reset_index(drop=True)
        st.dataframe(report_df)

        try:
            excel_data = create_excel_report(selected_item, stats, median_price, report_df)
            st.download_button(
                label="리포트 Excel로 다운로드",
                data=excel_data,
                file_name=f"{selected_item}_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Excel 파일 생성 중 오류가 발생했습니다: {e}")

# --------------------------------------------------------------------------
#                           5. 메인 실행 로직 (탭 구성)
# --------------------------------------------------------------------------

# --- 탭 생성 ---
tab1, tab2 = st.tabs(["공공 데이터 분석", "웹사이트 비급여 항목 분석"]) 

with tab1:
    public_df = load_data("data.parquet")
    if public_df is not None:
        # DataFrame에 해시 가능한 속성 추가
        public_df.hashCode = "public"
        create_analysis_page(public_df)
    else:
        st.error("'data.parquet' 파일을 찾을 수 없습니다. 앱과 같은 폴더에 파일을 넣어주세요.")

with tab2:
    crawled_df = load_data("crawled_data.parquet")
    if crawled_df is not None:
        # DataFrame에 해시 가능한 속성 추가
        crawled_df.hashCode = "crawled"
        create_analysis_page(crawled_df)
    else:
        st.warning("크롤링 데이터 파일('crawled_data.parquet')을 찾을 수 없습니다.")
        st.info("프로젝트 폴더에 해당 파일을 추가하면 기능이 활성화됩니다.")

# --- 사용자 피드백 (사이드바에 항상 표시) ---
st.sidebar.divider()
st.sidebar.header("사용자 피드백")
feedback_text = st.sidebar.text_area("앱에 대한 의견이나 개선점을 남겨주세요.", placeholder="피드백을 입력하세요...", key="feedback_text")
submit_button = st.sidebar.button("피드백 제출", key="feedback_submit")

if submit_button:
    if feedback_text:
        file_exists = os.path.isfile('feedback.csv')
        with open('feedback.csv', 'a', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(['timestamp', 'feedback'])
            writer.writerow([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), feedback_text])
        st.sidebar.success("소중한 의견 감사합니다! 피드백이 성공적으로 제출되었습니다.")
    else:
        st.sidebar.warning("피드백 내용이 비어있습니다. 내용을 입력해주세요.")
